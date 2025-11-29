import openpyxl
import os

def test_process_file():
    """测试处理tmp.xlsx文件"""
    file_path = 'tmp.xlsx'
    sheet_name = 'Sheet1'
    output_path = '.'
    
    print(f"开始测试处理文件: {file_path}")
    print(f"工作表: {sheet_name}")
    
    try:
        # 打开工作簿
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]
        
        # 创建新工作簿和工作表
        new_workbook = openpyxl.Workbook()
        new_sheet = new_workbook.active
        new_sheet.title = "漏洞详情处理结果"
        
        # 定义表头
        headers = ["序号", "漏洞标题", "漏洞编号", "漏洞类型", "危险级别", "影响平台", "CVSS分值", 
                  "bugtraq编号", "CVE编号", "CNCVE编号", "国家漏洞库编号", "CNNVD编号", 
                  "CNVD编号", "漏洞可利用性", "存在主机", "简单描述", "详细描述", "修补建议", "参考网址", "漏洞安全性"]
        new_sheet.append(headers)
        
        # 设置列宽
        for col in range(1, len(headers) + 1):
            new_sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
        
        # 遍历原始工作表，提取漏洞信息
        vulnerabilities = []
        current_vuln = {}
        vuln_index = 0
        
        for row in sheet.iter_rows(min_row=1, values_only=True):
            # 跳过空行
            if not any(row):
                continue
            
            print(f"处理行: {row}")
            
            # 检查是否是新漏洞标题行（以【数字】开头，在B列或A列）
            title_cell = row[1] if len(row) > 1 else row[0]
            if title_cell and isinstance(title_cell, str) and title_cell.startswith("【") and "】" in title_cell:
                # 如果有当前漏洞，先保存
                if current_vuln:
                    vulnerabilities.append(current_vuln)
                # 开始新漏洞
                current_vuln = {
                    "序号": title_cell.split("】")[0][1:],
                    "漏洞标题": title_cell.split("】")[1].strip()
                }
                vuln_index += 1
                print(f"发现新漏洞: {current_vuln['漏洞标题']}")
            # 检查是否是属性行（B列或A列有属性名称）
            elif len(row) >= 3 and row[1] and isinstance(row[1], str) and row[2]:
                # 提取属性名称和值（B列是属性名，C列是属性值）
                attr_name = row[1].strip()
                attr_value = row[2].strip()
                
                # 映射属性名称到表头
                attr_map = {
                    "漏洞编号": "漏洞编号",
                    "漏洞类型": "漏洞类型",
                    "危险级别": "危险级别",
                    "影响平台": "影响平台",
                    "CVSS分值": "CVSS分值",
                    "bugtraq编号": "bugtraq编号",
                    "CVE编号": "CVE编号",
                    "CNCVE编号": "CNCVE编号",
                    "国家漏洞库编号": "国家漏洞库编号",
                    "CNNVD编号": "CNNVD编号",
                    "CNVD编号": "CNVD编号",
                    "漏洞可利用性": "漏洞可利用性",
                    "存在主机": "存在主机",
                    "简单描述": "简单描述",
                    "详细描述": "详细描述",
                    "修补建议": "修补建议",
                    "参考网址": "参考网址",
                    "漏洞安全性": "漏洞安全性"
                }
                
                if attr_name in attr_map:
                    current_vuln[attr_map[attr_name]] = attr_value
                    print(f"  提取属性: {attr_name} = {attr_value}")
            # 兼容旧格式：A列是属性名，B列是属性值
            elif row[0] and isinstance(row[0], str) and len(row) > 1 and row[1]:
                attr_name = row[0].strip()
                attr_value = row[1].strip()
                
                attr_map = {
                    "漏洞编号": "漏洞编号",
                    "漏洞类型": "漏洞类型",
                    "危险级别": "危险级别",
                    "影响平台": "影响平台",
                    "CVSS分值": "CVSS分值",
                    "bugtraq编号": "bugtraq编号",
                    "CVE编号": "CVE编号",
                    "CNCVE编号": "CNCVE编号",
                    "国家漏洞库编号": "国家漏洞库编号",
                    "CNNVD编号": "CNNVD编号",
                    "CNVD编号": "CNVD编号",
                    "漏洞可利用性": "漏洞可利用性",
                    "存在主机": "存在主机",
                    "简单描述": "简单描述",
                    "详细描述": "详细描述",
                    "修补建议": "修补建议",
                    "参考网址": "参考网址",
                    "漏洞安全性": "漏洞安全性"
                }
                
                if attr_name in attr_map:
                    current_vuln[attr_map[attr_name]] = attr_value
                    print(f"  提取属性(旧格式): {attr_name} = {attr_value}")
        
        # 保存最后一个漏洞
        if current_vuln:
            vulnerabilities.append(current_vuln)
        
        print(f"\n成功提取 {len(vulnerabilities)} 个漏洞信息")
        
        # 将提取的漏洞信息写入新工作表
        for i, vuln in enumerate(vulnerabilities):
            print(f"\n写入漏洞 {i+1}: {vuln['漏洞标题']}")
            # 按照表头顺序提取值
            row_data = [vuln.get(header, "") for header in headers]
            new_sheet.append(row_data)
            print(f"  写入数据: {row_data[:5]}...")  # 只显示前5个字段
        
        # 保存新文件
        output_file = os.path.join(output_path, f"test_result_{os.path.basename(file_path)}")
        new_workbook.save(output_file)
        
        workbook.close()
        new_workbook.close()
        
        print(f"\n测试完成！结果保存至: {output_file}")
        return True
        
    except Exception as e:
        print(f"\n测试失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    test_process_file()
