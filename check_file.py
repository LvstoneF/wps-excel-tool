import openpyxl

# 打开Excel文件
wb = openpyxl.load_workbook('tmp.xlsx')

# 打印工作表名称
print('工作表名称:', wb.sheetnames)

# 遍历所有工作表
for sheet_name in wb.sheetnames:
    print(f'\n=== 工作表: {sheet_name} ===')
    sheet = wb[sheet_name]
    
    # 打印前20行内容
    print('前20行内容:')
    for row in sheet.iter_rows(min_row=1, max_row=20, values_only=True):
        # 过滤掉空行
        if any(row):
            print(row)

# 关闭工作簿
wb.close()
