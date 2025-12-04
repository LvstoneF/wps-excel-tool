import os
import re
import openpyxl
import xlrd
from constants import IP_PATTERN, IP_COLUMN_INDEX

class IPReplacer:
    """IP替换器，负责替换文件中的IP地址为设备名称"""
    
    def __init__(self, logger=None):
        """初始化IP替换器
        
        Args:
            logger (callable, optional): 日志记录函数，默认None
        """
        self.logger = logger
    
    def log(self, message):
        """记录日志
        
        Args:
            message (str): 日志消息
        """
        if self.logger:
            self.logger(message)
    
    def replace_ip_with_device(self, input_file, output_file, ip_device_map, ip_column_index=None):
        """替换Excel文件中的IP为设备名称，支持.xlsx和.xls格式
        
        Args:
            input_file (str): 输入文件路径
            output_file (str): 输出文件路径
            ip_device_map (dict): IP设备映射字典
            ip_column_index (int, optional): IP地址所在的列索引（从0开始），默认使用常量中的IP_COLUMN_INDEX
        
        Returns:
            str: 输出文件路径
        
        Raises:
            Exception: 替换IP失败时抛出异常
        """
        try:
            self.log(f"开始替换IP为设备名称: {input_file}")
            
            # 使用指定的列索引或默认值
            column_index = ip_column_index if ip_column_index is not None else IP_COLUMN_INDEX
            
            # 获取文件扩展名
            ext = os.path.splitext(input_file)[1].lower()
            
            if ext == '.xlsx':
                # 使用openpyxl处理.xlsx文件
                workbook = openpyxl.load_workbook(input_file)
                sheet = workbook.active
                
                # 遍历所有行，从第2行开始（跳过表头）
                replaced_count = 0
                for row in range(2, sheet.max_row + 1):
                    # 获取单元格值，Excel列从1开始
                    cell_value = sheet.cell(row=row, column=column_index + 1).value
                    if cell_value and isinstance(cell_value, str):
                        # 查找所有IP地址
                        ips = re.findall(IP_PATTERN, cell_value)
                        if ips:
                            # 替换每个IP地址
                            modified_value = cell_value
                            for ip in ips:
                                if ip in ip_device_map:
                                    device_name = ip_device_map[ip]
                                    # 只在设备名称有效（非空且不是"/"）的情况下进行替换
                                    if device_name and device_name != "/" and device_name != ip:
                                        modified_value = modified_value.replace(ip, device_name)
                                        replaced_count += 1
                                        self.log(f"  替换IP {ip} 为 {device_name}")
                                    else:
                                        self.log(f"  跳过替换IP {ip}，设备名称无效: {device_name}")
                            # 如果有修改，更新单元格值
                            if modified_value != cell_value:
                                sheet.cell(row=row, column=column_index + 1).value = modified_value
                
                # 保存输出文件
                workbook.save(output_file)
                workbook.close()
                
            elif ext == '.xls':
                # 使用xlrd读取.xls文件，然后使用openpyxl写入新的.xlsx文件
                self.log(f"检测到.xls文件，正在转换处理: {input_file}")
                
                # 读取.xls文件
                xlrd_workbook = xlrd.open_workbook(input_file)
                xlrd_sheet = xlrd_workbook.sheet_by_index(0)
                
                # 创建新的.xlsx文件
                openpyxl_workbook = openpyxl.Workbook()
                openpyxl_sheet = openpyxl_workbook.active
                
                # 复制表头
                header_row = xlrd_sheet.row_values(0)
                openpyxl_sheet.append(header_row)
                
                # 遍历所有行，从第2行开始（跳过表头）
                replaced_count = 0
                for row_idx in range(1, xlrd_sheet.nrows):
                    row_values = xlrd_sheet.row_values(row_idx)
                    
                    # 处理IP列
                    if len(row_values) > column_index:
                        cell_value = row_values[column_index]
                        if cell_value and isinstance(cell_value, str):
                            # 查找所有IP地址
                            ips = re.findall(IP_PATTERN, cell_value)
                            if ips:
                                # 替换每个IP地址
                                modified_value = cell_value
                                for ip in ips:
                                    if ip in ip_device_map:
                                        device_name = ip_device_map[ip]
                                        # 只在设备名称有效（非空且不是"/"）的情况下进行替换
                                        if device_name and device_name != "/" and device_name != ip:
                                            modified_value = modified_value.replace(ip, device_name)
                                            replaced_count += 1
                                            self.log(f"  替换IP {ip} 为 {device_name}")
                                        else:
                                            self.log(f"  跳过替换IP {ip}，设备名称无效: {device_name}")
                                # 如果有修改，更新单元格值
                                if modified_value != cell_value:
                                    row_values[column_index] = modified_value
                    
                    # 将处理后的行添加到新工作表
                    openpyxl_sheet.append(row_values)
                
                # 保存输出文件
                openpyxl_workbook.save(output_file)
                openpyxl_workbook.close()
                
            else:
                raise Exception(f"不支持的文件格式: {ext}")
            
            self.log(f"IP替换完成！共替换 {replaced_count} 个IP地址")
            self.log(f"替换结果保存至: {output_file}")
            return output_file
        except Exception as e:
            self.log(f"替换IP失败: {str(e)}")
            raise Exception(f"替换IP失败: {str(e)}")
