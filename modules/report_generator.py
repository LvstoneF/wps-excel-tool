import os
import openpyxl
from constants import HOST_STAT_SHEET_NAME

class ReportGenerator:
    """报告生成器，负责生成漏洞统计报告"""
    
    def __init__(self, logger=None):
        """初始化报告生成器
        
        Args:
            logger (callable, optional): 日志记录函数，默认None
        """
        self.logger = logger
    
    def log(self, message):
        """记录日志
        
        Args:
            message (str): 日志消息
        """
        if self.logger:
            self.logger(message)
    
    def merge_and_save_results(self, file_path, sheet_names, results, output_path):
        """合并多个工作表的处理结果并保存
        
        Args:
            file_path (str): 源文件路径
            sheet_names (list): 工作表名称列表
            results (list): 处理结果列表
            output_path (str): 输出目录路径
        
        Returns:
            str: 输出文件路径
        
        Raises:
            Exception: 合并结果失败时抛出异常
        """
        try:
            # 创建新工作簿和工作表
            new_workbook = openpyxl.Workbook()
            new_sheet = new_workbook.active
            new_sheet.title = "合并漏洞详情处理结果"
            
            # 定义表头
            headers = ["序号", "安全漏洞名称", "关联资产/域名", "严重程度"]
            new_sheet.append(headers)
            
            # 设置列宽
            column_widths = [10, 50, 20, 10]
            for i, width in enumerate(column_widths):
                new_sheet.column_dimensions[openpyxl.utils.get_column_letter(i+1)].width = width
            
            # 合并所有结果
            total_rows = 0
            for i, (sheet_name, result_data) in enumerate(zip(sheet_names, results)):
                self.log(f"合并{sheet_name}的处理结果，共{len(result_data)}行")
                # 写入数据（跳过表头，因为已经添加过了）
                for row in result_data:
                    new_sheet.append(row)
                    total_rows += 1
            
            # 保存新文件，确保使用.xlsx扩展名
            base_name = os.path.basename(file_path)
            name_without_ext = os.path.splitext(base_name)[0]
            output_file = os.path.join(output_path, f"合并处理结果_{name_without_ext}.xlsx")
            new_workbook.save(output_file)
            
            self.log(f"合并完成！共处理{len(sheet_names)}个工作表，生成{total_rows}行数据")
            self.log(f"结果保存至: {output_file}")
            return output_file
        except Exception as e:
            raise Exception(f"合并结果失败: {str(e)}")
    
    def generate_host_vuln_stat_report(self, file_path, hosts, vuln_counts, ip_device_map, output_path):
        """生成主机漏洞统计报告
        
        Args:
            file_path (str): 源文件路径
            hosts (list): 主机信息列表
            vuln_counts (dict): 漏洞统计字典
            ip_device_map (dict): IP设备映射字典
            output_path (str): 输出目录路径
        
        Returns:
            str: 输出文件路径
        
        Raises:
            Exception: 生成主机漏洞统计报告失败时抛出异常
        """
        try:
            self.log("开始生成主机漏洞统计报告")
            
            # 创建新工作簿和工作表
            new_workbook = openpyxl.Workbook()
            new_sheet = new_workbook.active
            new_sheet.title = HOST_STAT_SHEET_NAME
            
            # 定义表头
            headers = [
                "序号", 
                "设备名称或IP地址", 
                "系统及版本", 
                "安全漏洞数量", "", "", ""
            ]
            new_sheet.append(headers)
            
            # 定义二级表头
            sub_headers = [
                "", "", "", 
                "高", "中", "低", "小计"
            ]
            new_sheet.append(sub_headers)
            
            # 合并单元格
            new_sheet.merge_cells(start_row=1, start_column=4, end_row=1, end_column=7)
            
            # 设置列宽
            column_widths = [10, 30, 20, 10, 10, 10, 10]
            for i, width in enumerate(column_widths):
                new_sheet.column_dimensions[openpyxl.utils.get_column_letter(i+1)].width = width
            
            # 准备统计数据
            stat_data = []
            
            # 首先创建主机IP到信息的映射
            host_info_map = {}
            for host in hosts:
                # 查找IP地址
                ip = None
                for key, value in host.items():
                    import re
                    from constants import IP_PATTERN
                    if re.match(IP_PATTERN, value):
                        ip = value
                        break
                
                if ip:
                    # 获取系统及版本
                    os_version = ""
                    for key, value in host.items():
                        if "系统" in key or "版本" in key or "OS" in key:
                            os_version = value
                            break
                    
                    host_info_map[ip] = os_version
            
            # 按照映射表中设备出现的顺序进行排序
            # 首先处理映射表中的设备
            processed_ips = set()
            for ip, device_name in ip_device_map.items():
                # 获取系统及版本
                os_version = host_info_map.get(ip, "")
                
                # 获取漏洞统计数据
                counts = vuln_counts.get(ip, {"高": 0, "中": 0, "低": 0})
                
                # 计算小计
                total = counts["高"] + counts["中"] + counts["低"]
                
                # 添加到统计数据
                stat_data.append({
                    "ip": ip,
                    "device_name": device_name,
                    "os_version": os_version,
                    "counts": counts,
                    "total": total
                })
                
                processed_ips.add(ip)
            
            # 处理主机列表中未在映射表中出现的设备
            for host in hosts:
                # 查找IP地址
                ip = None
                for key, value in host.items():
                    import re
                    from constants import IP_PATTERN
                    if re.match(IP_PATTERN, value):
                        ip = value
                        break
                
                if ip and ip not in processed_ips:
                    # 获取设备名称
                    device_name = ip
                    
                    # 获取系统及版本
                    os_version = host_info_map.get(ip, "")
                    
                    # 获取漏洞统计数据
                    counts = vuln_counts.get(ip, {"高": 0, "中": 0, "低": 0})
                    
                    # 计算小计
                    total = counts["高"] + counts["中"] + counts["低"]
                    
                    # 添加到统计数据
                    stat_data.append({
                        "ip": ip,
                        "device_name": device_name,
                        "os_version": os_version,
                        "counts": counts,
                        "total": total
                    })
                    
                    processed_ips.add(ip)
            
            # 写入统计数据
            for i, data in enumerate(stat_data, start=1):
                row = [
                    i,
                    data["device_name"],
                    data["os_version"],
                    data["counts"]["高"],
                    data["counts"]["中"],
                    data["counts"]["低"],
                    data["total"]
                ]
                new_sheet.append(row)
            
            # 保存新文件，确保使用.xlsx扩展名
            base_name = os.path.basename(file_path)
            name_without_ext = os.path.splitext(base_name)[0]
            output_file = os.path.join(output_path, f"主机漏洞统计_{name_without_ext}.xlsx")
            new_workbook.save(output_file)
            
            self.log(f"主机漏洞统计报告生成完成！共统计 {len(stat_data)} 台设备")
            self.log(f"统计报告保存至: {output_file}")
            return output_file
        except Exception as e:
            self.log(f"生成主机漏洞统计报告失败: {str(e)}")
            raise Exception(f"生成主机漏洞统计报告失败: {str(e)}")
