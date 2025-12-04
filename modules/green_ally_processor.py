import os
import re
import openpyxl
import xlrd
from xlrd import xldate_as_tuple
import datetime
from constants import SEVERITY_MAP, IP_PATTERN, IP_COLUMN_INDEX

class GreenAllyProcessor:
    """绿盟漏扫报告处理器，负责处理绿盟漏扫生成的Excel文件"""
    
    def __init__(self, logger=None):
        """初始化绿盟漏扫报告处理器
        
        Args:
            logger (callable, optional): 日志记录函数，默认None
        """
        self.logger = logger
        # 绿盟漏扫报告中"远程漏洞"子表的固定名称
        self.REMOTE_VULN_SHEET_NAME = "远程漏洞"
        # 输出文件格式模板
        self.OUTPUT_TEMPLATE_1 = "合并处理结果_{name}.xlsx"
        self.OUTPUT_TEMPLATE_2 = "主机漏洞统计_{name}.xlsx"
        self.OUTPUT_TEMPLATE_3 = "替换IP后_合并处理结果_{name}.xlsx"
        self.OUTPUT_TEMPLATE_4 = "按漏洞名称合并结果_{name}.xlsx"
    
    def log(self, message):
        """记录日志
        
        Args:
            message (str): 日志消息
        """
        if self.logger:
            self.logger(message)
    
    def _is_header_row(self, data):
        """检查数据是否是需要排除的表头行
        
        Args:
            data (dict or list): 数据，可能是字典或列表
        
        Returns:
            bool: 是否是表头行
        """
        try:
            # 检查是否是字典格式（来自_extract_vulnerabilities方法）
            if isinstance(data, dict):
                vuln_name = data.get("漏洞名称", "").strip()
                asset = data.get("关联资产/域名", "").strip()
                risk_level = data.get("风险等级", "").strip()
                
                # 检查是否包含表头行的特征
                if vuln_name == "漏洞名称" or asset == "14.81.13.63" or risk_level == "风险等级":
                    return True
            
            # 检查是否是列表格式（标准格式）
            elif isinstance(data, list) and len(data) >= 4:
                # 检查序号字段
                if data[0].strip() == "1" or data[0].strip() == "序号":
                    # 检查其他字段
                    if data[1].strip() == "漏洞名称" or data[1].strip() == "安全漏洞名称":
                        return True
                    if data[2].strip() == "14.81.13.63" or data[2].strip() == "关联资产/域名":
                        return True
                    if data[3].strip() == "风险等级" or data[3].strip() == "严重程度":
                        return True
        except Exception as e:
            self.log(f"检查表头行时出错: {str(e)}")
        
        return False
    
    def batch_process_folder(self, folder_path, output_path, mapping_file_path=None):
        """批量处理指定文件夹中的所有.xls格式漏扫文件
        
        Args:
            folder_path (str): 文件夹路径
            output_path (str): 输出路径
            mapping_file_path (str, optional): IP设备映射表文件路径，默认None
        
        Returns:
            dict: 处理结果，包含成功和失败的文件信息
        """
        try:
            self.log(f"开始批量处理绿盟漏扫报告，文件夹: {folder_path}")
            
            # 检查文件夹是否存在
            if not os.path.exists(folder_path):
                raise Exception(f"文件夹不存在: {folder_path}")
            
            # 获取文件夹中所有.xls文件
            xls_files = [f for f in os.listdir(folder_path) if f.lower().endswith('.xls')]
            if not xls_files:
                self.log(f"未找到.xls文件")
                return {"success": [], "failed": []}
            
            self.log(f"找到 {len(xls_files)} 个.xls文件")
            
            # 处理每个文件
            success_files = []
            failed_files = []
            all_results = []
            
            for xls_file in xls_files:
                file_path = os.path.join(folder_path, xls_file)
                try:
                    # 验证文件命名规则
                    if not self._validate_file_name(xls_file):
                        self.log(f"文件命名不符合规则: {xls_file}")
                        failed_files.append({"file": xls_file, "reason": "文件命名不符合规则"})
                        continue
                    
                    # 处理文件
                    self.log(f"开始处理文件: {xls_file}")
                    results = self.process_green_ally_report(file_path)
                    all_results.extend(results)
                    success_files.append(xls_file)
                    self.log(f"成功处理文件: {xls_file}")
                except Exception as e:
                    self.log(f"处理文件失败: {xls_file}, 原因: {str(e)}")
                    failed_files.append({"file": xls_file, "reason": str(e)})
            
            # 生成输出文件
            if all_results:
                # 生成合并处理结果
                base_name = os.path.basename(folder_path)
                merged_file = self._generate_merged_result(all_results, output_path, base_name)
                
                # 生成主机漏洞统计报告（支持IP映射表）
                stat_file = self._generate_vuln_stat_report(all_results, output_path, base_name)
                
                # 生成按漏洞名称合并的结果文件
                vuln_merged_file = self._generate_vulnerability_merged_result(all_results, output_path, base_name)
                
                # 生成替换IP后的结果（如果提供了映射表）
                replaced_file = None
                replaced_vuln_merged_file = None
                replaced_stat_file = None
                ip_device_map = None
                
                if mapping_file_path:
                    from .ip_device_mapper import IPDeviceMapper
                    from .ip_replacer import IPReplacer
                    
                    ip_mapper = IPDeviceMapper(logger=self.log)
                    ip_device_map = ip_mapper.read_ip_device_mapping(mapping_file_path)
                    
                    # 使用IP映射表重新生成主机漏洞统计报告
                    stat_file = self._generate_vuln_stat_report(all_results, output_path, base_name, ip_device_map)
                    
                    ip_replacer = IPReplacer(logger=self.log)
                    
                    # 替换合并结果文件中的IP
                    replaced_file = os.path.join(output_path, self.OUTPUT_TEMPLATE_3.format(name=base_name))
                    ip_replacer.replace_ip_with_device(merged_file, replaced_file, ip_device_map)
                    
                    # 替换按漏洞名称合并结果文件中的IP
                    replaced_vuln_merged_file = os.path.join(output_path, f"替换IP后_按漏洞名称合并结果_{base_name}.xlsx")
                    ip_replacer.replace_ip_with_device(vuln_merged_file, replaced_vuln_merged_file, ip_device_map)
                    
                    # 替换统计报告文件中的IP
                    replaced_stat_file = os.path.join(output_path, f"替换IP后_主机漏洞统计_{base_name}.xlsx")
                    ip_replacer.replace_ip_with_device(stat_file, replaced_stat_file, ip_device_map, ip_column_index=1)
            
            self.log(f"批量处理完成，成功: {len(success_files)}，失败: {len(failed_files)}")
            return {"success": success_files, "failed": failed_files}
        except Exception as e:
            self.log(f"批量处理失败: {str(e)}")
            raise Exception(f"批量处理失败: {str(e)}")
    
    def _validate_file_name(self, file_name):
        """验证文件命名是否符合漏扫目标IP地址的规则要求
        
        Args:
            file_name (str): 文件名
        
        Returns:
            bool: 是否符合规则
        """
        # 简单验证：文件名是否包含IP地址
        # 可以根据实际规则调整
        base_name = os.path.splitext(file_name)[0]
        return re.search(IP_PATTERN, base_name) is not None
    
    def process_green_ally_report(self, file_path):
        """处理单个绿盟漏扫报告
        
        Args:
            file_path (str): 文件路径
        
        Returns:
            list: 处理后的漏洞数据列表
        """
        try:
            # 打开Excel文件
            workbook = xlrd.open_workbook(file_path)
            
            # 查找"远程漏洞"子表
            sheet_names = workbook.sheet_names()
            if self.REMOTE_VULN_SHEET_NAME not in sheet_names:
                raise Exception(f"未找到{self.REMOTE_VULN_SHEET_NAME}子表")
            
            sheet = workbook.sheet_by_name(self.REMOTE_VULN_SHEET_NAME)
            self.log(f"找到{self.REMOTE_VULN_SHEET_NAME}子表，共{sheet.nrows}行，{sheet.ncols}列")
            
            # 解析表头
            header = self._parse_header(sheet)
            if not header:
                raise Exception("无法解析表头")
            
            # 提取漏洞数据
            vulnerabilities = self._extract_vulnerabilities(sheet, header, file_path)
            
            # 转换为标准格式
            result_data = self._convert_to_standard_format(vulnerabilities, file_path)
            
            return result_data
        except Exception as e:
            self.log(f"处理绿盟漏扫报告失败: {str(e)}")
            raise Exception(f"处理绿盟漏扫报告失败: {str(e)}")
    
    def _parse_header(self, sheet):
        """解析表头
        
        Args:
            sheet (xlrd.sheet.Sheet): 工作表对象
        
        Returns:
            dict: 表头映射，{column_name: column_index}
        """
        for i in range(sheet.nrows):
            row = sheet.row_values(i)
            # 查找包含漏洞相关关键字的行作为表头
            if any(cell and isinstance(cell, str) and ("漏洞名称" in cell or "风险等级" in cell or "漏洞描述" in cell) for cell in row):
                header = {}
                for j, cell in enumerate(row):
                    if cell and isinstance(cell, str):
                        header[cell.strip()] = j
                return header
        return {}
    
    def _extract_vulnerabilities(self, sheet, header, file_path):
        """提取漏洞数据
        
        Args:
            sheet (xlrd.sheet.Sheet): 工作表对象
            header (dict): 表头映射
            file_path (str): 文件路径
        
        Returns:
            list: 漏洞数据列表
        """
        vulnerabilities = []
        
        # 从表头行的下一行开始读取数据
        for i in range(1, sheet.nrows):
            row = sheet.row_values(i)
            if not any(row):
                continue
            
            vuln = {}
            is_header_row = False
            
            # 提取关键字段
            if "漏洞名称" in header:
                vuln_name = str(row[header["漏洞名称"]]).strip() if row[header["漏洞名称"]] else ""
                vuln["漏洞名称"] = vuln_name
                # 检查是否是表头行
                if vuln_name == "漏洞名称":
                    is_header_row = True
            
            if "风险等级" in header:
                risk_level = str(row[header["风险等级"]]).strip() if row[header["风险等级"]] else ""
                vuln["风险等级"] = risk_level
                # 检查是否是表头行
                if risk_level == "风险等级":
                    is_header_row = True
            
            if "存在主机" in header or "关联资产" in header:
                asset_key = "存在主机" if "存在主机" in header else "关联资产"
                asset = str(row[header[asset_key]]).strip() if row[header[asset_key]] else ""
                vuln["关联资产/域名"] = asset
                # 检查是否是表头行
                if asset == "14.81.13.63":
                    is_header_row = True
            
            # 检查是否是需要排除的表头行
            if is_header_row:
                continue
            
            if "漏洞描述" in header:
                vuln["漏洞描述"] = str(row[header["漏洞描述"]]).strip() if row[header["漏洞描述"]] else ""
            
            if "解决方案" in header:
                vuln["解决方案"] = str(row[header["解决方案"]]).strip() if row[header["解决方案"]] else ""
            
            # 如果没有关联资产字段，尝试从文件名中提取IP
            if "关联资产/域名" not in vuln or not vuln["关联资产/域名"]:
                file_name = os.path.basename(file_path)
                base_name = os.path.splitext(file_name)[0]
                ips = re.findall(IP_PATTERN, base_name)
                if ips:
                    vuln["关联资产/域名"] = ips[0]
            
            vulnerabilities.append(vuln)
        
        return vulnerabilities
    
    def _convert_to_standard_format(self, vulnerabilities, file_path):
        """转换为标准格式
        
        Args:
            vulnerabilities (list): 漏洞数据列表
            file_path (str): 文件路径
        
        Returns:
            list: 标准格式的漏洞数据列表
        """
        result_data = []
        
        for i, vuln in enumerate(vulnerabilities, start=1):
            # 检查是否是需要排除的表头行
            vuln_name = vuln.get("漏洞名称", "").strip()
            asset = vuln.get("关联资产/域名", "").strip()
            risk_level = vuln.get("风险等级", "").strip()
            
            # 排除特定的表头行：漏洞名称、14.81.13.63、风险等级
            if vuln_name == "漏洞名称" and asset == "14.81.13.63" and risk_level == "风险等级":
                continue
            
            # 执行文本替换操作：将所有"[低]"、"[中]"、"[高]"替换为"低"、"中"、"高"
            risk_level = risk_level.replace("[低]", "低").replace("[中]", "中").replace("[高]", "高")
            
            # 映射风险等级到严重程度
            severity = SEVERITY_MAP.get(risk_level, risk_level)
            
            # 只保留严重程度不是"信息"的条目，并确保严重程度格式统一
            if severity != "信息":
                # 执行文本替换操作：将所有"[低]"、"[中]"、"[高]"替换为"低"、"中"、"高"
                severity = severity.replace("[低]", "低").replace("[中]", "中").replace("[高]", "高")
                
                # 统一严重程度格式，确保只有"高"、"中"、"低"三种等级
                if severity in ["高危险", "高危"]:
                    formatted_severity = "高"
                elif severity in ["中危险", "中危"]:
                    formatted_severity = "中"
                elif severity in ["低危险", "低危"]:
                    formatted_severity = "低"
                else:
                    formatted_severity = SEVERITY_MAP.get(severity, severity)
                    # 再次执行替换，确保格式统一
                    formatted_severity = formatted_severity.replace("[低]", "低").replace("[中]", "中").replace("[高]", "高")
                
                row_data = [
                    str(i),
                    vuln_name,
                    asset,
                    formatted_severity
                ]
                result_data.append(row_data)
        
        # 进一步过滤，确保结果中没有表头行
        filtered_results = []
        for row in result_data:
            if len(row) >= 4:
                # 检查是否是表头行：序号为"1"，漏洞名称为"漏洞名称"，关联资产为"14.81.13.63"，严重程度为"风险等级"
                if row[0].strip() == "1" and row[1].strip() == "漏洞名称" and row[2].strip() == "14.81.13.63" and row[3].strip() == "风险等级":
                    continue
            filtered_results.append(row)
        
        # 按严重程度排序：高 -> 中 -> 低
        severity_order = {"高": 0, "中": 1, "低": 2}
        sorted_results = sorted(filtered_results, key=lambda x: severity_order.get(x[3], 3))
        
        # 重新生成序号
        for i, row in enumerate(sorted_results, start=1):
            row[0] = str(i)
        
        return sorted_results
    
    def _generate_merged_result(self, all_results, output_path, base_name):
        """生成合并处理结果
        
        Args:
            all_results (list): 所有漏洞数据列表
            output_path (str): 输出路径
            base_name (str): 基础文件名
        
        Returns:
            str: 输出文件路径
        """
        try:
            self.log("开始生成合并处理结果")
            
            # 创建新工作簿和工作表
            new_workbook = openpyxl.Workbook()
            new_sheet = new_workbook.active
            new_sheet.title = "合并漏洞详情处理结果"
            
            # 定义表头
            headers = ["序号", "安全漏洞名称", "关联资产/域名", "严重程度"]
            new_sheet.append(headers)
            
            # 设置列宽
            column_widths = [10, 50, 20, 10]
            for i, width in enumerate(column_widths):
                new_sheet.column_dimensions[openpyxl.utils.get_column_letter(i+1)].width = width
            
            # 处理数据：格式化严重程度并按高中低排序
            processed_results = []
            for row in all_results:
                if len(row) < 4:
                    continue
                
                vuln_name = row[1].strip()
                asset = row[2].strip()
                severity = row[3].strip()
                
                # 检查是否是需要排除的表头行
                if vuln_name == "漏洞名称" and asset == "14.81.13.63" and severity == "风险等级":
                    continue
                
                # 执行文本替换操作：将所有"[低]"、"[中]"、"[高]"替换为"低"、"中"、"高"
                severity = severity.replace("[低]", "低").replace("[中]", "中").replace("[高]", "高")
                
                # 统一严重程度格式
                if severity in ["高危险", "高危"]:
                    formatted_severity = "高"
                elif severity in ["中危险", "中危"]:
                    formatted_severity = "中"
                elif severity in ["低危险", "低危"]:
                    formatted_severity = "低"
                else:
                    formatted_severity = SEVERITY_MAP.get(severity, severity)
                    # 执行文本替换操作
                    formatted_severity = formatted_severity.replace("[低]", "低").replace("[中]", "中").replace("[高]", "高")
                
                processed_results.append([
                    row[0],
                    vuln_name,
                    asset,
                    formatted_severity
                ])
            
            # 按严重程度排序：高 -> 中 -> 低
            severity_order = {"高": 0, "中": 1, "低": 2}
            sorted_results = sorted(processed_results, key=lambda x: severity_order.get(x[3], 3))
            
            # 重新生成序号
            for i, row in enumerate(sorted_results, start=1):
                row[0] = str(i)
            
            # 写入排序后的数据
            for row in sorted_results:
                new_sheet.append(row)
            
            # 保存文件
            output_file = os.path.join(output_path, self.OUTPUT_TEMPLATE_1.format(name=base_name))
            new_workbook.save(output_file)
            
            self.log(f"合并处理结果保存至: {output_file}")
            return output_file
        except Exception as e:
            self.log(f"生成合并处理结果失败: {str(e)}")
            raise Exception(f"生成合并处理结果失败: {str(e)}")
    
    def _generate_vuln_stat_report(self, all_results, output_path, base_name, ip_device_map=None):
        """生成主机漏洞统计报告
        
        Args:
            all_results (list): 所有漏洞数据列表
            output_path (str): 输出路径
            base_name (str): 基础文件名
            ip_device_map (dict, optional): IP设备映射字典，默认None
        
        Returns:
            str: 输出文件路径
        """
        try:
            self.log("开始生成主机漏洞统计报告")
            
            # 统计漏洞数量，支持IP映射表
            vuln_counts = self._count_vulnerabilities_by_ip(all_results, ip_device_map)
            
            # 创建新工作簿和工作表
            new_workbook = openpyxl.Workbook()
            new_sheet = new_workbook.active
            new_sheet.title = "主机漏洞统计"
            
            # 定义表头
            headers = [
                "序号", 
                "设备名称或IP地址", 
                "系统及版本", 
                "安全漏洞数量", "", "", ""
            ]
            new_sheet.append(headers)
            
            # 定义二级表头
            sub_headers = [
                "", "", "", 
                "高", "中", "低", "小计"
            ]
            new_sheet.append(sub_headers)
            
            # 合并单元格
            new_sheet.merge_cells(start_row=1, start_column=4, end_row=1, end_column=7)
            
            # 设置列宽
            column_widths = [10, 30, 20, 10, 10, 10, 10]
            for i, width in enumerate(column_widths):
                new_sheet.column_dimensions[openpyxl.utils.get_column_letter(i+1)].width = width
            
            # 按漏洞数量排序（先按高风险数量，再按中风险数量，最后按低风险数量）
            sorted_vuln_counts = sorted(
                vuln_counts.items(),
                key=lambda x: (-x[1]["高"], -x[1]["中"], -x[1]["低"])
            )
            
            # 写入统计数据
            for i, (device_ip, counts) in enumerate(sorted_vuln_counts, start=1):
                total = counts["高"] + counts["中"] + counts["低"]
                row = [
                    i,
                    device_ip,
                    "",
                    counts["高"],
                    counts["中"],
                    counts["低"],
                    total
                ]
                new_sheet.append(row)
            
            # 保存文件
            output_file = os.path.join(output_path, self.OUTPUT_TEMPLATE_2.format(name=base_name))
            new_workbook.save(output_file)
            
            self.log(f"主机漏洞统计报告保存至: {output_file}")
            return output_file
        except Exception as e:
            self.log(f"生成主机漏洞统计报告失败: {str(e)}")
            raise Exception(f"生成主机漏洞统计报告失败: {str(e)}")
    
    def _count_vulnerabilities_by_ip(self, all_results, ip_device_map=None):
        """按IP地址统计不同严重程度的漏洞数量，支持IP映射表
        
        Args:
            all_results (list): 所有漏洞数据列表
            ip_device_map (dict, optional): IP设备映射字典，默认None
        
        Returns:
            dict: 漏洞统计字典
        """
        vuln_counts = {}
        
        for row in all_results:
            if len(row) < 4:
                continue
            
            vuln_name = row[1].strip()
            ip_text = row[2].strip()
            severity = row[3].strip()
            
            # 检查是否是需要排除的表头行
            if vuln_name == "漏洞名称" and ip_text == "14.81.13.63" and severity == "风险等级":
                continue
            
            # 执行文本替换操作：将所有"[低]"、"[中]"、"[高]"替换为"低"、"中"、"高"
            severity = severity.replace("[低]", "低").replace("[中]", "中").replace("[高]", "高")
            
            # 统一严重程度格式
            if severity in ["高危险", "高危"]:
                formatted_severity = "高"
            elif severity in ["中危险", "中危"]:
                formatted_severity = "中"
            elif severity in ["低危险", "低危"]:
                formatted_severity = "低"
            else:
                formatted_severity = SEVERITY_MAP.get(severity, severity)
                # 执行文本替换操作
                formatted_severity = formatted_severity.replace("[低]", "低").replace("[中]", "中").replace("[高]", "高")
            
            # 只统计高、中、低三个级别
            if formatted_severity not in ["高", "中", "低"]:
                continue
            
            # 从文本中提取所有IP地址
            ips = re.findall(IP_PATTERN, ip_text)
            
            for ip in ips:
                if ip:
                    # 如果提供了IP映射表，使用映射表中的设备名称作为统计键
                    # 否则使用IP地址作为统计键
                    stat_key = ip_device_map[ip] if (ip_device_map and ip in ip_device_map) else ip
                    
                    if stat_key not in vuln_counts:
                        vuln_counts[stat_key] = {"高": 0, "中": 0, "低": 0, "ip": ip}  # 保存原始IP
                    
                    vuln_counts[stat_key][formatted_severity] += 1
        
        return vuln_counts
    
    def _merge_results_by_vulnerability(self, all_results):
        """按漏洞名称合并资产信息
        
        Args:
            all_results (list): 所有漏洞数据列表
        
        Returns:
            list: 按漏洞名称合并后的结果列表
        """
        try:
            self.log("开始按漏洞名称合并资产信息")
            
            # 创建字典，按漏洞名称分组，存储关联资产和严重程度
            vuln_dict = {}
            
            for row in all_results:
                if len(row) < 4:
                    continue
                
                vuln_name = row[1].strip()
                asset = row[2].strip()
                severity = row[3].strip()
                
                # 执行文本替换操作：将所有"[低]"、"[中]"、"[高]"替换为"低"、"中"、"高"
                severity = severity.replace("[低]", "低").replace("[中]", "中").replace("[高]", "高")
                
                # 检查是否是需要排除的表头行
                if vuln_name == "漏洞名称" and asset == "14.81.13.63" and severity == "风险等级":
                    continue
                
                if vuln_name:
                    if vuln_name not in vuln_dict:
                        # 新漏洞，初始化记录
                        # 统一严重程度格式
                        if severity in ["高危险", "高危"]:
                            formatted_severity = "高"
                        elif severity in ["中危险", "中危"]:
                            formatted_severity = "中"
                        elif severity in ["低危险", "低危"]:
                            formatted_severity = "低"
                        else:
                            formatted_severity = SEVERITY_MAP.get(severity, severity)
                            # 执行文本替换操作
                            formatted_severity = formatted_severity.replace("[低]", "低").replace("[中]", "中").replace("[高]", "高")
                        
                        vuln_dict[vuln_name] = {
                            "assets": set(),  # 使用集合避免重复资产
                            "severity": formatted_severity
                        }
                    
                    # 添加资产到集合
                    vuln_dict[vuln_name]["assets"].add(asset)
            
            # 转换为标准格式并按严重程度排序
            merged_results = []
            for vuln_name, data in vuln_dict.items():
                # 将资产集合转换为逗号分隔的字符串
                combined_assets = ", ".join(sorted(data["assets"]))
                merged_results.append([
                    "",  # 序号稍后生成
                    vuln_name,
                    combined_assets,
                    data["severity"]
                ])
            
            # 按严重程度排序：高 -> 中 -> 低
            severity_order = {"高": 0, "中": 1, "低": 2}
            sorted_results = sorted(merged_results, key=lambda x: severity_order.get(x[3], 3))
            
            # 生成序号
            for i, row in enumerate(sorted_results, start=1):
                row[0] = str(i)
            
            self.log(f"按漏洞名称合并完成，共合并 {len(sorted_results)} 个漏洞")
            return sorted_results
        except Exception as e:
            self.log(f"按漏洞名称合并失败: {str(e)}")
            raise Exception(f"按漏洞名称合并失败: {str(e)}")
    
    def _generate_vulnerability_merged_result(self, all_results, output_path, base_name):
        """生成按漏洞名称合并的结果文件
        
        Args:
            all_results (list): 所有漏洞数据列表
            output_path (str): 输出路径
            base_name (str): 基础文件名
        
        Returns:
            str: 输出文件路径
        """
        try:
            self.log("开始生成按漏洞名称合并的结果文件")
            
            # 按漏洞名称合并结果
            merged_results = self._merge_results_by_vulnerability(all_results)
            
            # 创建新工作簿和工作表
            new_workbook = openpyxl.Workbook()
            new_sheet = new_workbook.active
            new_sheet.title = "按漏洞名称合并结果"
            
            # 定义表头
            headers = ["序号", "安全漏洞名称", "关联资产/域名", "严重程度"]
            new_sheet.append(headers)
            
            # 设置列宽
            column_widths = [10, 50, 50, 10]  # 关联资产列宽度增加，以容纳多个资产
            for i, width in enumerate(column_widths):
                new_sheet.column_dimensions[openpyxl.utils.get_column_letter(i+1)].width = width
            
            # 写入数据
            for row in merged_results:
                new_sheet.append(row)
            
            # 保存文件
            output_file = os.path.join(output_path, self.OUTPUT_TEMPLATE_4.format(name=base_name))
            new_workbook.save(output_file)
            
            self.log(f"按漏洞名称合并结果保存至: {output_file}")
            return output_file
        except Exception as e:
            self.log(f"生成按漏洞名称合并结果失败: {str(e)}")
            raise Exception(f"生成按漏洞名称合并结果失败: {str(e)}")
    
    def is_green_ally_report(self, file_path):
        """判断是否是绿盟漏扫报告
        
        Args:
            file_path (str): 文件路径
        
        Returns:
            bool: 是否是绿盟漏扫报告
        """
        try:
            # 检查文件扩展名
            ext = os.path.splitext(file_path)[1].lower()
            if ext != '.xls':
                return False
            
            # 打开文件，检查是否包含"远程漏洞"子表
            workbook = xlrd.open_workbook(file_path)
            sheet_names = workbook.sheet_names()
            
            return self.REMOTE_VULN_SHEET_NAME in sheet_names
        except Exception as e:
            self.log(f"判断绿盟漏扫报告失败: {str(e)}")
            return False
